import sys
from PyQt5 import QtWidgets
from PyQt5 import uic
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
import shutil
from PIL import Image
from PyQt5.QtCore import Qt,QThread,pyqtSignal
import time



class Form(QtWidgets.QDialog):
    file_name = pyqtSignal(str)
    def __init__(self, parent=None):
        QtWidgets.QDialog.__init__(self, parent)
        self.ui = uic.loadUi("1012_QT_UI_V0.1.ui")
        self.ui.pushButton_1.clicked.connect(self.startProgressBar)
        self.ui.show()

    def filename(self):
        a =  self.ui.lineEdit_1.text()
        return a

    def startProgressBar(self):  # 프로그래스바 시작
        print("startProgressBar 시작 ")
        self.thread = MyThread()  # MyThread 클레스 선언
        # MytThread에서 사용한 change_value라는 pyqtSignal을 받아서 연결해준다
        self.thread.change_value.connect(self.setProgressVal) # change_value값이날라오면 연결한다 setProgressvVal 함수에 전달한다
        self.thread.start() # 쓰레드 실행, 파일이름을 함수에 전달
        global xlsx_dir  # 전역변수 지정
        xlsx_dir = self.ui.lineEdit_1.text()  # 엑셀 파일 이름 변수에 저장
        # print(xlsx_dir)
        print("startProgressBar 실행 ")




    def setProgressVal(self,Val): # Val값을 받아서 Progressbar를 업데이트 한다
        self.ui.progressBar_1.setValue(Val) # setProgressVal 함수 선언 Val를 전달 받아서 progessbar의 Setvalu를 Val로 입력
        print("setProgressVal 실행 ")

class MyThread(QThread):  # Qthread  실제 실행하고자 하는 함수 !!!!
    change_value = pyqtSignal(int)  # 시크널을 change_value 라는 변수에 담아서 보내준다
    print("MyThread 실행")

    def __init__(self): # 폼 구성
        super().__init__()

    def run(self):
        print("doaction")
        # xlsx_dir = "test"
        print(xlsx_dir)
        try:
            file_name = load_workbook("./" + xlsx_dir + ".xlsx")
            worksheet = file_name._sheets[0]   # sheet name or sheet number or list of sheet numbers and names
            count = 0
            fail = 0
            for row in worksheet.iter_rows():
                count += 1
                img_path = row[0].value  # 파일 경로 설정
                if os.path.isfile(img_path):
                    dir_path = './' + str(row[1].value) + '_' + str(row[2].value)  # 저장 경로 생성
                    if not os.path.exists(dir_path):  # 저장 경로가 없으면 신규 생성
                        os.mkdir(dir_path)
                    name = os.path.split(row[0].value)  # 파일 경로 분리
                    name = os.path.splitext(name[1])  # 파일 확장자 분리
                    save_path = dir_path + '/' + str(name[0]) + str(name[1])  # 저장 경로 설정
                    shutil.copy2(img_path, save_path)  # 파일 경로에 있는 이미지를 해당 폴더로 이동
                else:
                    print("FAIL : ", img_path)
                    fail += 1
                print(count)
                max_row = worksheet.max_row
                percent = round(100 * (count / max_row))
                self.change_value.emit(percent)

            print("total iamges : {:5d}, fail images : {:5d}".format(count, fail))
            print("완료")
            self.ui.pushButton_1.setText('완료')
        except:
            print("에러")

if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    w = Form()
    sys.exit(app.exec())